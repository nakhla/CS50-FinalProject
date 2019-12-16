from django.http import HttpResponse , JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.core import serializers
from .models import *
from datetime import datetime , timedelta , time
import json
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from openpyxl import Workbook

# Create your views here.
def index(request):
    
      return HttpResponseRedirect("/booking/")

def getResrvedLocationsByDayDict(requestedDay):
      
      date_time_obj = datetime.strptime(requestedDay, '%Y-%m-%d').date()
      loctimedateDictionary = {}
      currentDicIndex = 0
      currentVenueSlotDicIndex = 0
      venues = Venue.objects.all() #TODO: filter by active
      if len(venues) > 0:
            loctimedateDictionary['length'] = len(venues)
            for venue in venues:
                  loctimedateDictionary[currentDicIndex] = {'location' : venue.name, 'venueId': venue.id , 'venueSlots' : {}}
                  venueSlots = venue.venueDates.filter(day=date_time_obj).order_by('start_timeslot')
                  loctimedateDictionary[currentDicIndex]['venueSlots']['length'] = len(venueSlots)
                  if len(venueSlots) > 0:
                        for venueSlot in venueSlots:
                              loctimedateDictionary[currentDicIndex]['venueSlots'][currentVenueSlotDicIndex] = {
                                                                              'start_timeslot' : venueSlot.start_timeslot,
                                                                              'reserved_timeslots' : venueSlot.reserved_timeslots,
                                                                              'eventName' : venueSlot.event.name,
                                                                              'eventPk' : venueSlot.event.id}
                              currentVenueSlotDicIndex += 1
                                                                        
                  currentVenueSlotDicIndex = 0       
                  currentDicIndex += 1
      return loctimedateDictionary


def getResrvedLocationsByDay(request):

      requestedDay = request.GET["day"]
      loctimedateDictionary = getResrvedLocationsByDayDict(requestedDay)

      #locationsAndLocTimeFilteredByDay = Venue.objects.all().filter(venueDates__day=date_time_obj)
      #loctimedate = Loctimedate.objects.select_related('location').filter(day=date_time_obj) #TODO: Sort by Venue location then start time
      #loctimedate_serialized  = serializers.serialize('json', loctimedate)
      #return JsonResponse(loctimedate_serialized, safe=False)
      return JsonResponse(loctimedateDictionary, safe=False)


def insertReservedSlotsToEvent(request):
      requestedDay = request.GET["day"]
      eventId = request.GET["eventId"]
      Dict = json.loads(request.GET["cellsNeededToBeReserved"])

      formatedRequestedDay = datetime.strptime(requestedDay, '%Y-%m-%d').date()

      #TODO: make validations
      if len(Dict) > 0:
            for reserveObject in Dict:
                  start_timeslot = int(reserveObject["start_timeslot"])
                  reserved_timeslots = int(reserveObject["reserved_timeslots"])
                  # calculate starttime and endtime and write them to database
                  baseStartTime = time(8, 0, 0, 0)
                  startTime = baseStartTime.replace(hour=(baseStartTime.hour + start_timeslot) % 24)
                  endTime = startTime.replace(hour=(startTime.hour + reserved_timeslots) % 24)

                  Loctimedate.objects.create(event_id=int(eventId), day=formatedRequestedDay, location_id=reserveObject["venueId"], start_time=startTime, end_time=endTime, start_timeslot=start_timeslot, reserved_timeslots=reserved_timeslots)


      loctimedateDictionary = getResrvedLocationsByDayDict(requestedDay)

      return JsonResponse(loctimedateDictionary, safe=False)

@login_required
def booking(request):
      context = {
            'lastTenEvents':  Event.objects.all().order_by('-id')[:10],
            'categories':  Category.objects.all(),
            'organizers':  Organizer.objects.all(),
      }
      return render(request, "events/booking.html", context)

@login_required
def createEvent(request):

      eventName = request.GET["event-name"]
      eventDetails = request.GET["event-details"]
      eventCert = request.GET["event-cert"]
      eventcategory_id = request.GET["event-category"]
      eventorganizer_id = request.GET["event-organizer"]

      createdEvent = Event.objects.create(name=eventName, details=eventDetails, certificate=eventCert , user_id=request.user.id, category_id=eventcategory_id , organizer_id=eventorganizer_id )
 
      lastCreatedEvent = Event.objects.filter(id=createdEvent.id).all()
      lastCreatedEvent_serialized  = serializers.serialize('json', lastCreatedEvent)

      return JsonResponse(lastCreatedEvent_serialized, safe=False)

@login_required
def attendees(request, event_id):
      context = {
            'event': Event.objects.filter(id=event_id).get(),
            'attendees': Attendee.objects.filter(event=event_id).all()
      }
      return render(request,'events/attendees.html', context)

def addAttendee(request):
      firstName = request.GET["first_name"]
      lastName = request.GET["last_name"]
      nid = request.GET["nid"]
      email = request.GET["email"]
      mobile = request.GET["mobile"]
      event_id = request.GET["event_id"]
      retrievedAttendee = Attendee.objects.filter(nid=nid).all()
      if len(retrievedAttendee) == 0: #attendee not exists then create it
            attendee = Attendee.objects.create(first_name=firstName, last_name=lastName, nid=nid , email=email, mobile=mobile)
            attendee.save()
      attendee = Attendee.objects.filter(nid=nid).get()
      event = Event.objects.filter(id=event_id).get()
      event.attendees.add(attendee)
      event.save()
      lastCreatedAttendee = Attendee.objects.filter(id=attendee.id).all()
      lastCreatedAttendee_serialized  = serializers.serialize('json', lastCreatedAttendee)
      return JsonResponse(lastCreatedAttendee_serialized, safe=False)

def checkifAttExists(request):
      nid = request.GET["nid"]
      event_id = request.GET["event_id"]
      retrievedAttendee = Attendee.objects.filter(nid=nid).all()
      event = Event.objects.filter(id=event_id).get()
      if len(retrievedAttendee) != 0: # Existing Attendee
            attendeeInEvent = event.attendees.filter(id=retrievedAttendee[0].id)
            if len (attendeeInEvent) > 0: # this Attendee is existing in this Event Attendees
                  # Return ERror msg to be alerted Client Side
                  return JsonResponse({'ErrorMSG' : 'Attendee is already in this event!'}, safe=False)
            else: #Attendee is New for that Event
                  retrievedAttendee_serialized = serializers.serialize('json', retrievedAttendee)
                  return JsonResponse(retrievedAttendee_serialized, safe=False)
      
      return JsonResponse({'NewAttendee' : True}, safe=False)


def delAttendee(request):
      attendee_id = request.GET["attendee_id"]
      event_id = request.GET["event_id"]
      attendee = Attendee.objects.filter(id=attendee_id).get()
      event = Event.objects.filter(id=event_id).get()
      event.attendees.remove(attendee)
      event.save()
      return JsonResponse({'Deleted' : True}, safe=False)

def editAttendee(request):
      attendee_id = request.GET["attendee_id"]
      attendee = Attendee.objects.filter(id=attendee_id).all()
      attendee_serialized = serializers.serialize('json', attendee)
      return JsonResponse(attendee_serialized, safe=False)

def updateAttendee(request):
      
      att_id = int(request.POST["edit_attendeeid"])
      event_id = int(request.POST["edit_event_id"])
      att_nid =request.POST["edit_nid"]
      att_fn = request.POST["edit_first_name"]
      att_ln = request.POST["edit_last_name"]
      att_eml = request.POST["edit_email"]
      att_mob = request.POST["edit_mobile"]

      context = {
            'event': Event.objects.filter(id=event_id).get(),
            'attendees': Attendee.objects.filter(event=event_id).all()
      }

      attendee = Attendee.objects.filter(id=att_id).get()
      attendee.nid = att_nid
      attendee.first_name = att_fn
      attendee.last_name = att_ln
      attendee.email = att_eml
      attendee.mobile = att_mob
      attendee.save()
      messages.success(request, f"Attendee {att_id} - {att_fn} {att_ln} has been modified successfully")
      return render(request,'events/attendees.html', context)
      
    
@login_required
def events(request):
      context = {
            'lastTenEvents':  Event.objects.all().order_by('-id')[:10],
            'categories':  Category.objects.all(),
            'organizers':  Organizer.objects.all(),
            'loctimedates':  Loctimedate.objects.all(),
      }
      return render(request, "events/events.html", context)

@login_required
def attendeesList(request):
      context = {
            'attendees':  Attendee.objects.all(),
      }
      return render(request, "events/attendeesList.html", context)

def modifyAttendee(request):
      
      att_id = int(request.POST["edit_attendeeid"])
      
      att_nid =request.POST["edit_nid"]
      att_fn = request.POST["edit_first_name"]
      att_ln = request.POST["edit_last_name"]
      att_eml = request.POST["edit_email"]
      att_mob = request.POST["edit_mobile"]

      context = {
            'attendees': Attendee.objects.all()
      }

      attendee = Attendee.objects.filter(id=att_id).get()
      attendee.nid = att_nid
      attendee.first_name = att_fn
      attendee.last_name = att_ln
      attendee.email = att_eml
      attendee.mobile = att_mob
      attendee.save()
      messages.success(request, f"Attendee {att_id} - {att_fn} {att_ln} has been modified successfully")
      return render(request,"events/attendeesList.html", context)


@login_required
def filterEvents(request):

      cat_id = int(request.GET["category"])
      

      context = {
            
            'categories':  Category.objects.all(),
            'organizers':  Organizer.objects.all(),
            'loctimedates':  Loctimedate.objects.filter(event__category_id=cat_id).all(),
            'catSelected':  cat_id,
      }
      return render(request, "events/events.html", context)

def export_attendees_to_xlsx(request, event_id):
      """
      Downloads all attendees attending this event id as Excel file with a single worksheet
      """
      attendees_queryset = Attendee.objects.filter(event=event_id).all()
      event = Event.objects.filter(id=event_id).get()
      
      response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
      )
      response['Content-Disposition'] = 'attachment; filename={event}-{date}-attendees.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),event=event.name,
      )
      workbook = Workbook()
      
      # Get active worksheet/tab
      worksheet = workbook.active
      worksheet.title = event.name

      # Define the titles for columns
      columns = [
            'ID',
            'National ID',
            'First Name',
            'Last Name',
            'Email',
            'Mobile No',
      ]
      row_num = 1

      # Assign the titles for each cell of the header
      for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

      # Iterate through all attendees
      for attendee in attendees_queryset:
            row_num += 1
            
            # Define the data for each cell in the row 
            row = [
                  attendee.pk,
                  attendee.nid,
                  attendee.first_name,
                  attendee.last_name,
                  attendee.email,
                  attendee.mobile,
            ]
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                  cell = worksheet.cell(row=row_num, column=col_num)
                  cell.value = cell_value

      workbook.save(response)

      return response